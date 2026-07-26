"""Trade order email attachment and buyer-copy delivery."""

from __future__ import annotations

import base64
import io
import unittest
from datetime import datetime
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from ygo_app.email import _build_trade_order_body, send_trade_order_request
from ygo_app.trade_export import TRADE_ORDER_HEADERS, write_trade_order_xlsx


def _sample_lines():
    return [
        {
            "card_name": "Relinquished",
            "set_code": "SS01-ENC08",
            "set_name": "Starter Deck",
            "rarity_display": "C",
            "condition": "NearMint",
            "quantity": 1,
            "list_price": 0.31,
            "offer_price": None,
            "comment": "Ez egy comment",
        }
    ]


class TestTradeOrderEmail(unittest.TestCase):
    def test_write_trade_order_xlsx_headers_and_row(self):
        content = write_trade_order_xlsx(_sample_lines())
        wb = load_workbook(io.BytesIO(content))
        headers = [cell.value for cell in next(wb.active.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, TRADE_ORDER_HEADERS)
        row = next(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(
            row,
            (
                "Relinquished",
                "SS01-ENC08",
                "Starter Deck",
                "C",
                "NearMint",
                1,
                0.31,
                None,
                "Ez egy comment",
            ),
        )

    def test_body_summarizes_items_without_inline_card_details(self):
        subject, body = _build_trade_order_body(
            seller_display_name="Owner Shop",
            buyer_contact={"name": "Buyer", "email": "buyer@test.example"},
            lines=_sample_lines(),
            submitted_at=datetime(2026, 7, 26, 22, 2, 36),
        )
        self.assertEqual(subject, "Trade order request — Owner Shop")
        self.assertIn("See attached Excel file for 1 item.", body)
        self.assertNotIn("Relinquished", body)
        self.assertNotIn("SS01-ENC08", body)
        self.assertIn("Email: buyer@test.example", body)

    @patch("ygo_app.email.EMAIL_FROM", "YGO Collection <noreply@example.com>")
    @patch("ygo_app.email.BREVO_API_KEY", "test-key")
    @patch("ygo_app.email.EMAIL_BACKEND", "brevo")
    @patch("ygo_app.email.requests.post")
    def test_brevo_seller_payload_includes_attachment(self, post_mock):
        post_mock.return_value = MagicMock(status_code=201, text="ok")
        send_trade_order_request(
            owner_email="owner@test.example",
            seller_display_name="Owner Shop",
            buyer_contact={"name": "Buyer", "email": "buyer@test.example"},
            lines=_sample_lines(),
            submitted_at=datetime(2026, 7, 26, 22, 2, 36),
            send_copy_to_buyer=False,
        )
        self.assertEqual(post_mock.call_count, 1)
        payload = post_mock.call_args.kwargs["json"]
        self.assertEqual(payload["to"], [{"email": "owner@test.example", "name": "Owner Shop"}])
        self.assertEqual(payload["replyTo"]["email"], "buyer@test.example")
        self.assertEqual(len(payload["attachment"]), 1)
        self.assertEqual(payload["attachment"][0]["name"], "trade-order.xlsx")
        raw = base64.b64decode(payload["attachment"][0]["content"])
        wb = load_workbook(io.BytesIO(raw))
        self.assertEqual(wb.active["A2"].value, "Relinquished")

    @patch("ygo_app.email.EMAIL_FROM", "YGO Collection <noreply@example.com>")
    @patch("ygo_app.email.BREVO_API_KEY", "test-key")
    @patch("ygo_app.email.EMAIL_BACKEND", "brevo")
    @patch("ygo_app.email.requests.post")
    def test_brevo_sends_separate_buyer_copy(self, post_mock):
        post_mock.return_value = MagicMock(status_code=201, text="ok")
        send_trade_order_request(
            owner_email="owner@test.example",
            seller_display_name="Owner Shop",
            buyer_contact={"name": "Buyer", "email": "buyer@test.example"},
            lines=_sample_lines(),
            submitted_at=datetime(2026, 7, 26, 22, 2, 36),
            send_copy_to_buyer=True,
        )
        self.assertEqual(post_mock.call_count, 2)
        seller_payload = post_mock.call_args_list[0].kwargs["json"]
        buyer_payload = post_mock.call_args_list[1].kwargs["json"]
        self.assertEqual(seller_payload["to"][0]["email"], "owner@test.example")
        self.assertEqual(buyer_payload["to"], [{"email": "buyer@test.example", "name": "Buyer"}])
        self.assertEqual(buyer_payload["replyTo"]["email"], "owner@test.example")
        self.assertNotIn("owner@test.example", [r["email"] for r in buyer_payload["to"]])
        self.assertEqual(
            seller_payload["attachment"][0]["content"],
            buyer_payload["attachment"][0]["content"],
        )


if __name__ == "__main__":
    unittest.main()
